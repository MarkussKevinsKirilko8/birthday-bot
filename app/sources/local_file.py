from __future__ import annotations
import openpyxl
from app.models import Employee
from app.sources.base import COLUMNS, row_to_employee


class LocalFileSource:
    def __init__(self, path: str, sheet: str = "Sheet1"):
        self.path = path
        self.sheet = sheet

    def load(self) -> list[Employee]:
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = wb[self.sheet] if self.sheet in wb.sheetnames else wb.worksheets[0]
        header = [c.value for c in ws[1]]
        out: list[Employee] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            record = {header[i]: row[i] for i in range(len(header)) if header[i] in COLUMNS}
            emp = row_to_employee(record)
            if emp:
                out.append(emp)
        return out
